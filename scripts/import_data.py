#!/usr/bin/env python3
"""
One-time migration script: imports data from the original Excel file
into the new flat-structure attendance.xlsx.
"""

import os
import sys
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import load_workbook
from services.excel_service import (
    _ensure_file, add_team_member, add_holiday, set_balance,
    _read_sheet, _write_sheet, SHEETS,
)

ORIGINAL_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "Holidays and sick leaves - Counting.xlsx",
)

ACTIVE_MEMBERS = {"Danilo", "Ottavio", "Daniele", "Andrei", "Davide"}

MONTH_MAP = {
    "Gennaio": 1, "Febbraio": 2, "Marzo": 3, "Aprile": 4,
    "Maggio": 5, "Giugno": 6, "Luglio": 7, "Agosto": 8,
    "Setembre": 9, "Settembre": 9, "Ottobre": 10, "Novembre": 11, "Dicembre": 12,
}

YEARS = [22, 23, 24, 25, 26]


def parse_year(suffix):
    """Convert 2-digit suffix to 4-digit year."""
    return 2000 + suffix


def import_team():
    """Import active team members."""
    print("Importing team members...")
    for name in sorted(ACTIVE_MEMBERS):
        try:
            add_team_member(name, 2022)
            print(f"  + {name}")
        except ValueError:
            print(f"  ~ {name} (already exists)")


def import_holidays(wb):
    """Import public holidays from Summary sheets."""
    print("Importing holidays...")
    count = 0
    for year_suffix in YEARS:
        year = parse_year(year_suffix)
        sheet_name = f"Summary {year_suffix}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, max_col=3, values_only=True):
            date_val, day_name, holiday_name = row
            if not holiday_name or not date_val:
                continue
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                continue
            try:
                add_holiday(date_str, str(holiday_name), year)
                count += 1
            except Exception:
                pass
    print(f"  Imported {count} holidays")


def import_balances(wb):
    """Import vacation balances from Summary sheets."""
    print("Importing balances...")
    count = 0
    for year_suffix in YEARS:
        year = parse_year(year_suffix)
        sheet_name = f"Summary {year_suffix}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, max_col=10, values_only=True):
            if len(row) < 10:
                continue
            name = row[5]  # Column F
            if not name or name not in ACTIVE_MEMBERS:
                continue
            carried = row[6] or 0    # Column G
            accrued = row[7] or 0    # Column H
            used = row[8] or 0       # Column I
            remaining = row[9] or 0  # Column J
            try:
                set_balance(
                    str(name), year,
                    round(float(carried), 2),
                    round(float(accrued), 2),
                    round(float(used), 2),
                    round(float(remaining), 2),
                )
                count += 1
            except Exception as e:
                print(f"  ! Error for {name} {year}: {e}")
    print(f"  Imported {count} balance records")


def import_leave(wb):
    """Import vacation and sick leave entries from individual sheets."""
    print("Importing leave entries...")
    vacation_records = []
    sick_records = []

    for year_suffix in YEARS:
        year = parse_year(year_suffix)
        for member_name in ACTIVE_MEMBERS:
            sheet_name = f"{member_name} {year_suffix}"
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=13, max_col=5, values_only=True):
                month_name = row[0]
                if not month_name or month_name == "Totale":
                    continue
                month_num = MONTH_MAP.get(str(month_name).strip())
                if not month_num:
                    continue

                # Vacation (columns B, C)
                vac_count = row[1]
                vac_dates = row[2]
                if vac_count is not None:
                    try:
                        vac_count = float(str(vac_count).replace(",", "."))
                    except (ValueError, TypeError):
                        vac_count = 0
                    vacation_records.append({
                        "name": member_name,
                        "year": year,
                        "month": month_num,
                        "days_count": vac_count,
                        "dates_detail": str(vac_dates) if vac_dates else "",
                    })

                # Sick leave (columns D, E)
                sick_count = row[3] if len(row) > 3 else None
                sick_dates = row[4] if len(row) > 4 else None
                if sick_count is not None:
                    try:
                        sick_count = float(str(sick_count).replace(",", "."))
                    except (ValueError, TypeError):
                        sick_count = 0
                    sick_records.append({
                        "name": member_name,
                        "year": year,
                        "month": month_num,
                        "days_count": sick_count,
                        "dates_detail": str(sick_dates) if sick_dates else "",
                    })

    # Sort and write
    vacation_records.sort(key=lambda r: (r["name"], r["year"], r["month"]))
    sick_records.sort(key=lambda r: (r["name"], r["year"], r["month"]))
    _write_sheet("vacation", vacation_records)
    _write_sheet("sick_leave", sick_records)
    print(f"  Imported {len(vacation_records)} vacation entries, {len(sick_records)} sick leave entries")


def main():
    if not os.path.exists(ORIGINAL_FILE):
        print(f"Error: Original file not found at {ORIGINAL_FILE}")
        sys.exit(1)

    print(f"Reading: {ORIGINAL_FILE}")
    wb = load_workbook(ORIGINAL_FILE, data_only=True)

    _ensure_file()
    import_team()
    import_holidays(wb)
    import_balances(wb)
    import_leave(wb)

    wb.close()
    print("\nMigration complete! Data written to data/attendance.xlsx")


if __name__ == "__main__":
    main()
