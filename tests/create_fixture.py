"""Generate a minimal anonymized fixture workbook for tests.

Run directly to regenerate: python tests/create_fixture.py

Layout mirrors the real 'Holidays and sick leaves - Counting.xlsx':
- Person sheets: "{Name} {YY}" with Italian month headers, Totale SUM row
- Summary sheets: "Summary {YY}" with holidays (cols A-C), balances (cols F-J)
"""

import os
from datetime import datetime
from openpyxl import Workbook

FIXTURE_PATH = os.path.join(os.path.dirname(__file__), "fixtures", "minimal.xlsx")

MONTHS_IT = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Setembre", "Ottobre", "Novembre", "Dicembre",
]

MEMBERS = ["Alice", "Bob"]

# Known leave data for deterministic balance tests.
# Format: {(name, year_suffix): {month_1based: (vacation_count, vacation_detail, sick_count, sick_detail)}}
LEAVE_DATA = {
    ("Alice", 25): {
        1: (2, "15, 16", None, None),
        3: (1, "10", None, None),
        6: (5, "2-6", None, None),
        7: (5, "14-18", None, None),
        9: (0.5, "12(1/2)", None, None),
        12: (3, "22, 23, 24", 1, "5"),
    },
    ("Alice", 26): {
        1: ("1,5", "2, 23(1/2)", None, None),  # Italian decimal string
        4: (2.5, "10(1/2), 16, 17", None, None),
    },
    ("Bob", 25): {
        2: (3, "3, 4, 5", None, None),
        8: (10, "4-8, 18-22", None, None),
        11: (1, "20", 2, "10, 11"),
    },
    ("Bob", 26): {
        3: (1, "5", None, None),
    },
}

# Holidays (subset — enough to test weekend-holiday counting)
HOLIDAYS = {
    25: [
        (datetime(2025, 1, 1), "Wed", "New Year's Day"),
        (datetime(2025, 4, 21), "Mon", "Easter Monday"),
        (datetime(2025, 5, 1), "Thu", "Labour Day"),
        (datetime(2025, 5, 9), "Fri", "Europe Day"),
        (datetime(2025, 5, 29), "Thu", "Ascension Day"),
        (datetime(2025, 6, 9), "Mon", "Whit Monday"),
        (datetime(2025, 6, 23), "Mon", "National Day"),
        (datetime(2025, 8, 15), "Fri", "Assumption Day"),
        (datetime(2025, 11, 1), "Sat", "All Saints' Day"),     # weekend
        (datetime(2025, 12, 25), "Thu", "Christmas Day"),
        (datetime(2025, 12, 26), "Fri", "2nd Day of Christmas"),
    ],
    26: [
        (datetime(2026, 1, 1), "Thu", "New Year's Day"),
        (datetime(2026, 4, 6), "Mon", "Easter Monday"),
        (datetime(2026, 5, 1), "Fri", "Labour Day"),
        (datetime(2026, 5, 9), "Sat", "Europe Day"),           # weekend
        (datetime(2026, 5, 14), "Thu", "Ascension Day"),
        (datetime(2026, 5, 25), "Mon", "Whit Monday"),
        (datetime(2026, 6, 23), "Tue", "National Day"),
        (datetime(2026, 8, 15), "Sat", "Assumption Day"),      # weekend
        (datetime(2026, 11, 1), "Sun", "All Saints' Day"),     # weekend
        (datetime(2026, 12, 25), "Fri", "Christmas Day"),
        (datetime(2026, 12, 26), "Sat", "2nd Day of Christmas"),  # weekend
    ],
}

# Weekend holidays count per year (manually verified from above)
WEEKEND_HOLIDAYS = {25: 1, 26: 4}


def _create_person_sheet(wb, name, year_suffix):
    """Create a person sheet with Italian headers and Totale SUM formulas."""
    sheet_name = f"{name} {year_suffix}"
    ws = wb.create_sheet(sheet_name)

    # Row 1: headers
    ws.cell(row=1, column=1, value="Mesi")
    ws.cell(row=1, column=2, value="Ferie")
    ws.cell(row=1, column=3, value="Giorni ferie")
    ws.cell(row=1, column=4, value="Malattia")
    ws.cell(row=1, column=5, value="Giorni malattia")

    # Rows 2-13: months
    for i, month_name in enumerate(MONTHS_IT):
        row = i + 2
        ws.cell(row=row, column=1, value=month_name)

    # Fill known leave data
    key = (name, year_suffix)
    if key in LEAVE_DATA:
        for month, (vac_count, vac_detail, sick_count, sick_detail) in LEAVE_DATA[key].items():
            row = month + 1
            ws.cell(row=row, column=2, value=vac_count)
            ws.cell(row=row, column=3, value=vac_detail)
            if sick_count is not None:
                ws.cell(row=row, column=4, value=sick_count)
            if sick_detail is not None:
                ws.cell(row=row, column=5, value=sick_detail)

    # Row 14: Totale with SUM formulas
    ws.cell(row=14, column=1, value="Totale")
    ws.cell(row=14, column=2, value="=SUM(B2:B13)")
    ws.cell(row=14, column=4, value="=SUM(D2:D13)")


def _create_summary_sheet(wb, year_suffix, weekend_count):
    """Create a Summary sheet with holidays and balance formulas."""
    sheet_name = f"Summary {year_suffix}"
    ws = wb.create_sheet(sheet_name)

    full_year = 2000 + year_suffix

    # Row 3: headers
    ws.cell(row=3, column=1, value="Date")
    ws.cell(row=3, column=2, value="Day")
    ws.cell(row=3, column=3, value="Holiday")
    ws.cell(row=3, column=6, value="Name")
    ws.cell(row=3, column=7, value="Vacation days remaining from the year before")
    ws.cell(row=3, column=8, value="Vacation days accrued this year")
    ws.cell(row=3, column=9, value="Vacation days used this year")
    ws.cell(row=3, column=10, value="Total vacation to use until March next year")

    # Holidays (cols A-C, rows 4+)
    holidays = HOLIDAYS.get(year_suffix, [])
    for i, (dt, day_name, hol_name) in enumerate(holidays):
        row = 4 + i
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=2, value=day_name)
        ws.cell(row=row, column=3, value=hol_name)

    # Row 16: weekend holidays count
    ws.cell(row=16, column=1, value="# holidays in the weekend:")
    ws.cell(row=16, column=3, value=weekend_count)

    # Row 17: total info
    ws.cell(row=17, column=1, value=f"Total vacation days: 26 (+{weekend_count})")

    # Row 19-20: accrual rate
    ws.cell(row=19, column=1, value="Vacation days accrued each month: ")
    ws.cell(row=20, column=1, value="= 26/12")

    # Members balance section (cols F-J, rows 4+)
    for i, name in enumerate(MEMBERS):
        row = 4 + i
        ws.cell(row=row, column=6, value=name)

        # Col G: carried over — for the oldest year use 0, otherwise reference previous year
        if year_suffix == 25:
            ws.cell(row=row, column=7, value=0)
        else:
            # Simplified formula: reference previous Summary's total column
            # Real file uses structured table refs; we use a simpler A20-style formula
            # that the code can parse, plus the literal fallback path
            prev_suffix = year_suffix - 1
            ws.cell(row=row, column=7,
                    value=f"='{name} {prev_suffix}'!J{row}")  # formula string (won't be parsed by code — triggers literal/recursive path)

        # Col H: accrued — use the formula the code's regex expects
        ws.cell(row=row, column=8,
                value=f"=ROUNDUP($A$20*12+$C$16, 1)")

        # Col I: used — reference person sheet's Totale
        ws.cell(row=row, column=9,
                value=f"='{name} {year_suffix}'!B14")

        # Col J: total = carried + accrued - used
        g = ws.cell(row=row, column=7).coordinate
        h = ws.cell(row=row, column=8).coordinate
        ii = ws.cell(row=row, column=9).coordinate
        ws.cell(row=row, column=10,
                value=f"={g}+{h}-{ii}")


def create_fixture():
    """Build and save the minimal fixture workbook."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for year_suffix in [26, 25]:  # newest first, like real file
        _create_summary_sheet(wb, year_suffix, WEEKEND_HOLIDAYS[year_suffix])
        for name in MEMBERS:
            _create_person_sheet(wb, name, year_suffix)

    os.makedirs(os.path.dirname(FIXTURE_PATH), exist_ok=True)
    wb.save(FIXTURE_PATH)
    print(f"Fixture saved to {FIXTURE_PATH}")
    wb.close()


if __name__ == "__main__":
    create_fixture()
