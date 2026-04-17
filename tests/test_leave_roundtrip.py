"""Leave CRUD round-trip tests — verify Excel survives writes correctly."""

import openpyxl

from services.excel_service import (
    add_leave,
    delete_leave,
    get_leave,
    update_leave,
)


def test_add_then_get_vacation(patch_data_file):
    """Write a vacation entry, read it back, assert values match."""
    result = add_leave("vacation", "Alice", 2026, 5, 2, "11, 12")
    assert result["days_count"] == 2
    assert result["dates_detail"] == "11, 12"

    entries = get_leave("vacation", year=2026, name="Alice", month=5)
    assert len(entries) == 1
    assert entries[0]["days_count"] == 2
    assert entries[0]["dates_detail"] == "11, 12"


def test_add_merges_existing_month(patch_data_file):
    """Adding to a month that already has data merges counts and details."""
    # Alice 26 month 4 already has 2.5 days / "10(1/2), 16, 17"
    result = add_leave("vacation", "Alice", 2026, 4, 1, "25")
    assert result["days_count"] == 3.5
    assert "10(1/2), 16, 17" in result["dates_detail"]
    assert "25" in result["dates_detail"]


def test_add_sick_leave(patch_data_file):
    result = add_leave("sick_leave", "Bob", 2026, 2, 1, "10")
    assert result["days_count"] == 1

    entries = get_leave("sick_leave", year=2026, name="Bob", month=2)
    assert len(entries) == 1
    assert entries[0]["days_count"] == 1


def test_totale_sum_formula_preserved_after_add(patch_data_file):
    """The SUM formula in row 14 must never be overwritten to a value."""
    add_leave("vacation", "Alice", 2026, 5, 2, "11, 12")

    wb = openpyxl.load_workbook(patch_data_file)
    ws = wb["Alice 26"]

    # Col B row 14: vacation total
    assert ws.cell(row=14, column=2).value == "=SUM(B2:B13)"
    # Col D row 14: sick total
    assert ws.cell(row=14, column=4).value == "=SUM(D2:D13)"
    wb.close()


def test_update_leave(patch_data_file):
    """Overwrite an existing leave entry."""
    update_leave("vacation", "Alice", 2026, 1, {
        "days_count": 3,
        "dates_detail": "2, 23, 24",
    })
    entries = get_leave("vacation", year=2026, name="Alice", month=1)
    assert entries[0]["days_count"] == 3
    assert entries[0]["dates_detail"] == "2, 23, 24"


def test_delete_leave_clears_cells(patch_data_file):
    """Delete clears both count and detail to None."""
    delete_leave("vacation", "Alice", 2026, 1)

    entries = get_leave("vacation", year=2026, name="Alice", month=1)
    assert len(entries) == 0

    # Verify cells are actually None
    wb = openpyxl.load_workbook(patch_data_file)
    ws = wb["Alice 26"]
    assert ws.cell(row=2, column=2).value is None  # month 1 = row 2
    assert ws.cell(row=2, column=3).value is None
    wb.close()


def test_formula_survives_update_and_delete(patch_data_file):
    """Totale SUM formula intact after update + delete cycle."""
    update_leave("vacation", "Bob", 2026, 3, {"days_count": 5, "dates_detail": "1-5"})
    delete_leave("vacation", "Bob", 2026, 3)

    wb = openpyxl.load_workbook(patch_data_file)
    ws = wb["Bob 26"]
    assert ws.cell(row=14, column=2).value == "=SUM(B2:B13)"
    wb.close()
