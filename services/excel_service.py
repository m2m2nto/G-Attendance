import json
import os
import threading
from datetime import datetime
from openpyxl import load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
_DEFAULT_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "Holidays and sick leaves - Counting.xlsx",
)
_SETTINGS_FILE = os.path.join(DATA_DIR, "app_settings.json")

_lock = threading.Lock()

# Italian month names used in the Excel file (rows 2-13 of person sheets)
_MONTH_NAMES_IT = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Setembre", "Ottobre", "Novembre", "Dicembre",
]

# Person sheet layout:
#   Row 1:  Headers (Mesi, Ferie, Giorni ferie, Malattia, Giorni malattia)
#   Row 2:  Gennaio   | vacation_count | vacation_detail | sick_count | sick_detail
#   ...
#   Row 13: Dicembre
#   Row 14: Totale (SUM formulas — NEVER overwrite)
#
# Summary sheet layout (rows start at 3):
#   Row 3:  Headers
#   Row 4+: Col A-C = holidays, Col F-J = team balances (with formulas)


# ---------------------------------------------------------------------------
# File path management
# ---------------------------------------------------------------------------

def _load_settings():
    if os.path.exists(_SETTINGS_FILE):
        with open(_SETTINGS_FILE, "r") as f:
            return json.load(f)
    return {}


def _save_settings(settings):
    os.makedirs(os.path.dirname(_SETTINGS_FILE), exist_ok=True)
    with open(_SETTINGS_FILE, "w") as f:
        json.dump(settings, f, indent=2)


def get_data_file():
    settings = _load_settings()
    return settings.get("data_file", _DEFAULT_FILE)


def set_data_file(path):
    if not path:
        raise ValueError("File path cannot be empty")
    if not os.path.isfile(path):
        raise FileNotFoundError(f"File not found: {path}")
    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("File must be an Excel file (.xlsx or .xlsm)")
    try:
        wb = load_workbook(path, read_only=True)
        wb.close()
    except Exception as e:
        raise ValueError(f"Cannot open Excel file: {e}")
    _check_file_not_open(path)
    settings = _load_settings()
    settings["data_file"] = path
    _save_settings(settings)
    return path


def _check_file_not_open(filepath=None):
    """Check if the Excel file is open by another application."""
    if filepath is None:
        filepath = get_data_file()
    if not os.path.exists(filepath):
        return
    dirname = os.path.dirname(filepath)
    basename = os.path.basename(filepath)
    lock_file = os.path.join(dirname, "~$" + basename)
    if os.path.exists(lock_file):
        raise IOError(
            f"The file '{basename}' appears to be open in another application. "
            "Please close it and try again."
        )
    try:
        with open(filepath, "r+b"):
            pass
    except PermissionError:
        raise IOError(
            f"The file '{basename}' is locked by another application. "
            "Please close it and try again."
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _year_suffix(year):
    """Convert a 4-digit year to the 2-digit suffix used in sheet names."""
    return str(year)[-2:]


def _available_years():
    """Return sorted list of 4-digit years found in the workbook."""
    wb = load_workbook(get_data_file(), read_only=True)
    years = set()
    for name in wb.sheetnames:
        if name.startswith("Summary "):
            suffix = name.split(" ", 1)[1]
            try:
                yy = int(suffix)
                years.add(2000 + yy if yy < 100 else yy)
            except ValueError:
                pass
    wb.close()
    return sorted(years)


def _parse_days_count(val):
    """Parse a days count that might be a number, Italian-format string, or None."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("", "-", " - ", "  - "):
        return 0
    # Italian decimal comma: "1,5" -> 1.5
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0


def _open_for_read():
    """Open the workbook in read-only mode with computed values."""
    return load_workbook(get_data_file(), read_only=True, data_only=True)


def _open_for_write():
    """Open the workbook for writing (preserves formulas and VBA if .xlsm)."""
    _check_file_not_open()
    path = get_data_file()
    keep_vba = path.lower().endswith(".xlsm")
    return load_workbook(path, keep_vba=keep_vba)


def _save(wb):
    """Save and close the workbook."""
    wb.save(get_data_file())
    wb.close()


# ---------------------------------------------------------------------------
# Config (stored in app_settings.json, not in the Excel file)
# ---------------------------------------------------------------------------

def get_config():
    settings = _load_settings()
    return {
        "base_vacation_days": settings.get("base_vacation_days", "26"),
        "accrual_per_month": settings.get("accrual_per_month", "2.1667"),
        "carryover_deadline_month": settings.get("carryover_deadline_month", "3"),
    }


def update_config(key, value):
    settings = _load_settings()
    settings[key] = value
    _save_settings(settings)
    return get_config()


# ---------------------------------------------------------------------------
# Team — derived from person sheets in the workbook
# ---------------------------------------------------------------------------

def get_team():
    """Build team list from sheet names. A member is 'active' if they have a sheet
    for the most recent year."""
    wb = load_workbook(get_data_file(), read_only=True)
    years = _available_years()
    latest_year = years[-1] if years else datetime.now().year

    members = {}  # name -> {start_year, end_year}
    for name in wb.sheetnames:
        if name.startswith("Summary "):
            continue
        parts = name.rsplit(" ", 1)
        if len(parts) != 2:
            continue
        person, suffix = parts
        try:
            yy = int(suffix)
            year = 2000 + yy if yy < 100 else yy
        except ValueError:
            continue
        if person not in members:
            members[person] = {"name": person, "start_year": year, "end_year": None}
        else:
            if year < members[person]["start_year"]:
                members[person]["start_year"] = year
    wb.close()

    # Mark members without a sheet for the latest year as "left"
    latest_suffix = _year_suffix(latest_year)
    wb2 = load_workbook(get_data_file(), read_only=True)
    sheet_names = set(wb2.sheetnames)
    wb2.close()
    for person, info in members.items():
        if f"{person} {latest_suffix}" not in sheet_names:
            # Find latest year they have a sheet for
            for y in reversed(years):
                if f"{person} {_year_suffix(y)}" in sheet_names:
                    info["end_year"] = y
                    break

    return list(members.values())


def add_team_member(name, start_year):
    """Create sheets for a new team member for the given year."""
    suffix = _year_suffix(start_year)
    sheet_name = f"{name} {suffix}"
    with _lock:
        wb = _open_for_write()
        if sheet_name in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' already exists")
        ws = wb.create_sheet(sheet_name)
        # Headers
        ws.cell(row=1, column=1, value="Mesi")
        ws.cell(row=1, column=2, value="Ferie")
        ws.cell(row=1, column=3, value="Giorni ferie")
        ws.cell(row=1, column=4, value="Malattia")
        ws.cell(row=1, column=5, value="Giorni malattia")
        # Month rows
        for i, month_name in enumerate(_MONTH_NAMES_IT):
            ws.cell(row=i + 2, column=1, value=month_name)
        # Totale row with SUM formulas
        ws.cell(row=14, column=1, value="Totale")
        ws.cell(row=14, column=2, value="=SUM(B2:B13)")
        ws.cell(row=14, column=4, value="=SUM(D2:D13)")
        _save(wb)
    return {"name": name, "start_year": start_year, "end_year": None}


def update_team_member(name, data):
    # Team data is derived from sheet names — no direct update needed
    # Just return current info
    team = get_team()
    for m in team:
        if m["name"] == name:
            return m
    raise ValueError(f"Team member '{name}' not found")


def delete_team_member(name):
    # We don't delete sheets — just mark as inactive by not having current year sheets
    pass


# ---------------------------------------------------------------------------
# Holidays — from Summary sheets
# ---------------------------------------------------------------------------

def get_holidays(year=None):
    wb = _open_for_read()
    results = []
    years_to_check = [year] if year else _available_years()
    for y in years_to_check:
        suffix = _year_suffix(y)
        sheet_name = f"Summary {suffix}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            date_val = row[0] if len(row) > 0 else None
            holiday_name = row[2] if len(row) > 2 else None
            if date_val is None or holiday_name is None:
                break
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)
            day_name = row[1] if len(row) > 1 else ""
            results.append({
                "date": date_str,
                "name": holiday_name,
                "day": day_name,
                "year": y,
            })
    wb.close()
    return results


def add_holiday(date, name, year):
    # Holidays are part of the Summary sheet structure — adding requires care
    suffix = _year_suffix(year)
    sheet_name = f"Summary {suffix}"
    with _lock:
        wb = _open_for_write()
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"No Summary sheet for year {year}")
        ws = wb[sheet_name]
        # Find first empty row in the holidays section (col A, starting row 4)
        insert_row = None
        for r in range(4, ws.max_row + 2):
            if ws.cell(row=r, column=1).value is None:
                insert_row = r
                break
        if insert_row is None:
            insert_row = ws.max_row + 1
        ws.cell(row=insert_row, column=1, value=date)
        ws.cell(row=insert_row, column=2, value="")  # Day name — will be set manually
        ws.cell(row=insert_row, column=3, value=name)
        _save(wb)
    return {"date": date, "name": name, "year": year}


def update_holiday(old_date, old_year, data):
    suffix = _year_suffix(old_year)
    sheet_name = f"Summary {suffix}"
    with _lock:
        wb = _open_for_write()
        ws = wb[sheet_name]
        for r in range(4, ws.max_row + 1):
            cell_date = ws.cell(row=r, column=1).value
            if cell_date is None:
                break
            if isinstance(cell_date, datetime):
                cell_date_str = cell_date.strftime("%Y-%m-%d")
            else:
                cell_date_str = str(cell_date)
            if cell_date_str == str(old_date):
                if "date" in data:
                    ws.cell(row=r, column=1, value=data["date"])
                if "name" in data:
                    ws.cell(row=r, column=3, value=data["name"])
                _save(wb)
                return data
        wb.close()
    raise ValueError("Holiday not found")


def delete_holiday(date, year):
    suffix = _year_suffix(year)
    sheet_name = f"Summary {suffix}"
    with _lock:
        wb = _open_for_write()
        ws = wb[sheet_name]
        # Find the row to delete
        delete_row = None
        last_holiday_row = 3  # header row
        for r in range(4, ws.max_row + 1):
            cell_date = ws.cell(row=r, column=1).value
            if cell_date is None:
                break
            last_holiday_row = r
            if isinstance(cell_date, datetime):
                cell_date_str = cell_date.strftime("%Y-%m-%d")
            else:
                cell_date_str = str(cell_date)
            if cell_date_str == str(date):
                delete_row = r

        if delete_row is None:
            wb.close()
            return

        # Shift subsequent holiday rows up to close the gap
        for r in range(delete_row, last_holiday_row):
            for col in range(1, 4):
                ws.cell(row=r, column=col, value=ws.cell(row=r + 1, column=col).value)
        # Clear the last row
        for col in range(1, 4):
            ws.cell(row=last_holiday_row, column=col, value=None)
        _save(wb)


# ---------------------------------------------------------------------------
# Leave (vacation & sick_leave) — from person sheets
# ---------------------------------------------------------------------------

def _get_leave_col(leave_type):
    """Return (count_col, detail_col) 1-based column indices."""
    if leave_type == "vacation":
        return 2, 3  # Col B, C
    else:  # sick_leave
        return 4, 5  # Col D, E


def get_leave(leave_type, year=None, name=None, month=None):
    """Read leave entries from person sheets."""
    wb = _open_for_read()
    count_col, detail_col = _get_leave_col(leave_type)
    results = []
    years_to_check = [year] if year else _available_years()

    for y in years_to_check:
        suffix = _year_suffix(y)
        # Determine which people to check
        if name:
            people = [name]
        else:
            people = [
                sn.rsplit(" ", 1)[0]
                for sn in wb.sheetnames
                if sn.endswith(f" {suffix}") and not sn.startswith("Summary")
            ]

        for person in people:
            sheet_name = f"{person} {suffix}"
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            months_to_check = [month] if month else range(1, 13)
            for m in months_to_check:
                row = m + 1  # Month 1 = row 2
                count_val = ws.cell(row=row, column=count_col).value
                detail_val = ws.cell(row=row, column=detail_col).value
                days = _parse_days_count(count_val)
                if days > 0 or detail_val:
                    results.append({
                        "name": person,
                        "year": y,
                        "month": m,
                        "days_count": days,
                        "dates_detail": str(detail_val) if detail_val else "",
                    })
    wb.close()
    return results


def add_leave(leave_type, name, year, month, days_count, dates_detail):
    """Write a leave entry to the person's sheet.
    If the month already has data, merge the new days into the existing entry.
    """
    suffix = _year_suffix(year)
    sheet_name = f"{name} {suffix}"
    count_col, detail_col = _get_leave_col(leave_type)
    row = month + 1  # Month 1 = row 2

    with _lock:
        wb = _open_for_write()
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"No sheet found for {name} in {year}")
        ws = wb[sheet_name]
        existing_count = _parse_days_count(ws.cell(row=row, column=count_col).value)
        existing_detail = ws.cell(row=row, column=detail_col).value

        new_count = existing_count + days_count
        if existing_detail and dates_detail:
            new_detail = f"{existing_detail}, {dates_detail}"
        else:
            new_detail = dates_detail or (str(existing_detail) if existing_detail else None)

        ws.cell(row=row, column=count_col, value=new_count)
        ws.cell(row=row, column=detail_col, value=new_detail)
        _save(wb)

    return {
        "name": name, "year": year, "month": month,
        "days_count": new_count, "dates_detail": new_detail or "",
    }


def update_leave(leave_type, name, year, month, data):
    """Update an existing leave entry in the person's sheet."""
    suffix = _year_suffix(year)
    sheet_name = f"{name} {suffix}"
    count_col, detail_col = _get_leave_col(leave_type)
    row = month + 1

    with _lock:
        wb = _open_for_write()
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"No sheet found for {name} in {year}")
        ws = wb[sheet_name]
        if "days_count" in data:
            ws.cell(row=row, column=count_col, value=data["days_count"])
        if "dates_detail" in data:
            ws.cell(row=row, column=detail_col, value=data["dates_detail"] or None)
        _save(wb)

    return {
        "name": name, "year": year, "month": month,
        "days_count": data.get("days_count", 0),
        "dates_detail": data.get("dates_detail", ""),
    }


def delete_leave(leave_type, name, year, month):
    """Clear a leave entry from the person's sheet."""
    suffix = _year_suffix(year)
    sheet_name = f"{name} {suffix}"
    count_col, detail_col = _get_leave_col(leave_type)
    row = month + 1

    with _lock:
        wb = _open_for_write()
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        ws = wb[sheet_name]
        ws.cell(row=row, column=count_col, value=None)
        ws.cell(row=row, column=detail_col, value=None)
        _save(wb)


# ---------------------------------------------------------------------------
# Balances — computed from Summary sheets
# ---------------------------------------------------------------------------

def _get_summary_members(year):
    """Read team member names from the Summary sheet (col F, rows 4+)."""
    suffix = _year_suffix(year)
    sheet_name = f"Summary {suffix}"
    wb = load_workbook(get_data_file(), read_only=True)
    names = []
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(4, ws.max_row + 1):
            val = ws.cell(row=r, column=6).value
            if val is None:
                break
            names.append(val)
    wb.close()
    return names


def _count_weekend_holidays(year):
    holidays = get_holidays(year=year)
    count = 0
    for h in holidays:
        date_str = h.get("date")
        if not date_str:
            continue
        try:
            d = datetime.strptime(str(date_str), "%Y-%m-%d")
            if d.weekday() >= 5:
                count += 1
        except (ValueError, TypeError):
            pass
    return count


def _parse_accrual_formula(formula_str):
    """Parse the accrual formula like '=ROUNDUP(A20*11+2, 1)' or '=ROUNDUP($A$20*12+$C$16, 1)'.

    Returns (months, weekend_holidays) or None if unparseable.
    The formula pattern is: ROUNDUP(A20 * months + weekend_count, 1)
    where A20 = 26/12 (base_days / 12).
    """
    import re
    if not formula_str or not isinstance(formula_str, str):
        return None
    # Normalize: remove =, spaces, dollar signs
    f = formula_str.replace("=", "").replace(" ", "").replace("$", "")
    # Match ROUNDUP(A20*X+Y,1) or ROUNDUP(A20*X+C16,1)
    m = re.match(r'ROUNDUP\(A20\*([0-9.]+)\+([0-9.]+|C16),\s*1\)', f, re.IGNORECASE)
    if m:
        months = float(m.group(1))
        wh_str = m.group(2)
        if wh_str.upper() == "C16":
            return months, None  # weekend holidays stored in C16, resolve separately
        return months, float(wh_str)
    return None


def _get_weekend_holidays_from_summary(year):
    """Read the weekend holidays count from the Summary sheet's C16 cell,
    or from the text in A17 ('Total vacation days: 26 (+N)').
    Falls back to counting from holiday dates.
    """
    import re
    suffix = _year_suffix(year)
    sheet_name = f"Summary {suffix}"
    wb = load_workbook(get_data_file(), read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return _count_weekend_holidays(year)
    ws = wb[sheet_name]
    # Try C16 first (some years have it)
    c16 = ws.cell(row=16, column=3).value
    if isinstance(c16, (int, float)):
        wb.close()
        return int(c16)
    # Try parsing A17 text: "Total vacation days: 26 (+4)"
    for r in range(15, 20):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str) and "Total vacation days" in val:
            m = re.search(r'\(\+(\d+)\)', val)
            if m:
                wb.close()
                return int(m.group(1))
    wb.close()
    return _count_weekend_holidays(year)


def _get_entitlement(name, year):
    """Compute the entitlement for a member/year by parsing the accrual formula
    from the Summary sheet. Falls back to base_days + weekend_holidays.
    """
    import math
    config = get_config()
    base_days = float(config.get("base_vacation_days", 26))
    accrual_monthly = base_days / 12
    weekend_holidays = _get_weekend_holidays_from_summary(year)

    suffix = _year_suffix(year)
    sheet_name = f"Summary {suffix}"
    wb = load_workbook(get_data_file(), read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return base_days + weekend_holidays

    ws = wb[sheet_name]
    for r in range(4, ws.max_row + 1):
        member_name = ws.cell(row=r, column=6).value
        if member_name is None:
            break
        if member_name == name:
            formula = ws.cell(row=r, column=8).value
            wb.close()
            parsed = _parse_accrual_formula(formula)
            if parsed:
                months, wh = parsed
                if wh is None:
                    wh = weekend_holidays
                return math.ceil((accrual_monthly * months + wh) * 10) / 10
            break
    else:
        wb.close()

    return base_days + weekend_holidays


def _get_carried_over(name, year):
    """Compute carried-over for a member/year.

    2022 (the oldest year) stores literal 0 in the Summary — read it directly.
    For later years, recursively compute: previous year's remaining.
    """
    suffix = _year_suffix(year)
    sheet_name = f"Summary {suffix}"

    # Check if the carried_over cell is a literal value (not a formula)
    wb = load_workbook(get_data_file(), read_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(4, ws.max_row + 1):
            member_name = ws.cell(row=r, column=6).value
            if member_name is None:
                break
            if member_name == name:
                cell_val = ws.cell(row=r, column=7).value
                wb.close()
                # If it's a plain number (not a formula string), use it directly
                if isinstance(cell_val, (int, float)):
                    return float(cell_val)
                # It's a formula — compute from previous year
                prev_years = sorted([y for y in _available_years() if y < year], reverse=True)
                if prev_years:
                    prev = compute_balance(name, prev_years[0])
                    return prev["remaining"]
                return 0
    wb.close()
    return 0


def compute_balance(name, year):
    """Compute vacation balance entirely from formulas + raw person-sheet data.

    - entitlement: parsed from the Summary accrual formula
    - carried_over: literal value or computed from previous year
    - used: summed from person sheet cells (handles Italian string values)
    """
    entitlement = _get_entitlement(name, year)
    carried = _get_carried_over(name, year)
    weekend_holidays = _get_weekend_holidays_from_summary(year)

    vacation_entries = get_leave("vacation", year=year, name=name)
    used = sum(e["days_count"] for e in vacation_entries)

    total = round(carried + entitlement, 2)
    remaining = round(total - used, 2)
    return {
        "name": name,
        "year": year,
        "carried_over": round(carried, 2),
        "entitlement": entitlement,
        "weekend_holidays": weekend_holidays,
        "used": used,
        "total": total,
        "remaining": remaining,
    }


def get_balances(year=None, name=None):
    """Compute balances for all members from raw data."""
    years_to_check = [year] if year else _available_years()
    results = []
    for y in years_to_check:
        members = [name] if name else _get_summary_members(y)
        for member in members:
            b = compute_balance(member, y)
            results.append({
                "name": member,
                "year": y,
                "carried_over": b["carried_over"],
                "accrued": b["entitlement"],
                "used": b["used"],
                "remaining": b["remaining"],
            })
    return results


def set_balance(name, year, carried_over, accrued, used, remaining):
    b = compute_balance(name, year)
    return {
        "name": name, "year": year,
        "carried_over": b["carried_over"], "accrued": b["entitlement"],
        "used": b["used"], "remaining": b["remaining"],
    }
